import cv2
import numpy as np
import matplotlib.pyplot as plt
import os
from openpyxl import Workbook

from typing import Dict, List, Optional, Tuple


def surround(array: np.ndarray, coord: Tuple[int, int]):
    '''Takes in an array and a pixel's coordinate. Returns an array that contains the pixel itself and its eight surrounding pixels, in total nine.
    '''
    lis = coord[0]
    ele = coord[1]
    surround = array[lis-1:lis+2, ele-1:ele+2]
    return surround


def area(image: np.ndarray):
    '''Takes in a labelled marker image. Returns a list that contains the area of each grain.
    '''
    # Get the numbers of pixels with each label
    label, area = np.unique(image, return_counts=True)

    data = list(zip(label, area))
    # Discard backgroud and boundary pixels
    data = data[2:]

    return data


def fore_back(image: np.ndarray):
    '''Takes in a labelled marker image. Returns the area fractions of the foreground and the background.
    Foreground are labelled with integers starting from 2, background with 1 and boundaries with -1.
    '''
    # Get total area of labelled regions
    area_back = np.count_nonzero(image == 1)
    # Get total area of the image
    area = image.shape[0] * image.shape[1]
    # Get total area of background
    area_fore = area - area_back
    
    return {
        'foreground area': area_fore,
        'background area': area_back,
        'total area': area,
        'foreground fraction': area_fore/area,
        'background fraction': area_back/area
    }


def circumference(image: np.ndarray, visualise=False):
    '''Takes in a labelled marker image. Returns a list that contains the circumference of each grain.
    '''
    # A blank image of the same dimension as the mock image. Used for labelling
    blank = np.zeros(
        image.shape, np.uint8)
    labels = np.unique(image)[2:]

    # Get a certain label (positive integer that represent one segmented region)
    for label_no in labels:
        result = np.where(image == label_no)  # Exclude other labels
        coordinates = list(zip(result[0], result[1]))
        for coord in coordinates:
            neighbours = surround(image, coord)
            # Test if the surrounding pixels contain -1, the bounary.
            if -1 in neighbours:
                # If so, put a mark on the blank image
                blank[coord] = label_no

    # Merge the data
    label, circumference = np.unique(blank, return_counts=True)
    circumference = circumference[2:]
    label = label[2:]
    data = list(zip(label, circumference))

    if visualise == False:
        return data
    else:
        return blank  # Visualisation of the boundaries


def width_length_ellipse(image: np.ndarray, label, visualise=False):
    '''Takes in a labelled marker image and a specific label. Returns the minor- and major-axis length of an ellipse that will fit the grain with that label.
    '''
    blank = np.zeros(
        image.shape, np.uint8)  # A blank image of the same dimension as the mock image. Used for extraction grain pixels of one label.
    result = np.where(image == label)  # Obtain the coordinates of each label

    coordinates = list(zip(result[0], result[1]))
    for coord in coordinates:
        blank[coord] = 1
    contours, hierarchy = cv2.findContours(
        blank, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
    cnt = contours[0]
    ellipse = cv2.fitEllipse(cnt)  # Get data about the geometry of the ellipse
    blank = cv2.ellipse(blank, ellipse, 2, 3)  # Visualise

    if visualise == False:
        return ellipse[1]
    else:
        return blank


def width_length_rectangle(image: np.ndarray, label, visualise=False):
    '''Takes in a labelled marker image and a specific label. Returns the width and length of a rectangle that will fit the grain with that label.
    '''
    blank = np.zeros(
        image.shape, np.uint8)  # A blank image of the same dimension as the mock image. Used for extraction grain pixels of one label.
    result = np.where(image == label)  # Obtain the coordinates of each label

    coordinates = list(zip(result[0], result[1]))
    for coord in coordinates:
        blank[coord] = 1
    contours, hierarchy = cv2.findContours(blank, 1, 2)
    cnt = contours[0]
    rect = cv2.minAreaRect(cnt)
    box = cv2.boxPoints(rect)
    box = np.int0(box)  # Get data about the geometry of the rectangle
    blank = cv2.drawContours(blank, [box], 0, 2, 3)  # Visualise

    if visualise == False:
        return rect[1]
    else:
        return blank


def width_length_size(image: np.ndarray, method='ellipse'):
    '''Takes in a labelled marker image. Returns a list containing tuples that record the grain label, width, length and average of each grain.
    '''
    unique = np.unique(image)[2:]  # Get all labels

    data_l = []

    for i in unique:
        try:
            # Use width_length_ellipse or width_length rectangle to generate data for one label
            if method == 'ellipse':
                data = width_length_ellipse(image, i)
                data_l.append((i, data[0], data[1], (data[0]+data[1])/2))
            elif method == 'rectangle':
                data = width_length_rectangle(image, i)
                data_l.append((i, data[0], data[1], (data[0]+data[1])/2))
            else:
                raise ValueError(
                    'method string must be either <ellipse> or <rectangle>')
        except:
            print('Error processing %d' % i)
            pass

    return data_l


def data_extraction(image: np.ndarray, filename):
    '''Takes in a labelled marker image. Saves a Microsoft Excel workbook that contains the grain data in the working directory.
    Also prints the alpha and beta phase fraction.
    '''
    wb = Workbook()
    ws = wb.active

    ar, circum, wl = (None, None, None)

    # Try to generate the.data
    try:
        ar = area(image)
    except:
        print('Unable to generate area data')
        pass
    try:
        circum = circumference(image)
    except:
        print('Unable to generate circumference data')
        pass
    try:
        wl = width_length_size(image)
    except:
        print('Unable to generate width, length and size data.')
        pass
    unique = np.unique(image)[2:]

    # Print total number of grains
    print('There are in total %d grains.' % len(unique))

    # Print alpha and beta volume fractions
    alpha_beta = fore_back(image)
    print('Alpha volume fraction: {}'.format(round(alpha_beta['foreground fraction'], 2)))
    print('Beta volume fraction: {}'.format(round(alpha_beta['background fraction'], 2)))

    # Ensure that area, circumference and siwidth&length data match. If not, discard circumference or width&length data.
    if len(circum) == len(ar):
        pass
    else:
        print('Area and circumference data do not match. Discarding circumference data.')
        circum = None
    if len(wl) == len(ar):
        pass
    else:
        print('Area and width&length data do not match. Discarding width&length data.')
        wl = None

    # Create headers
    ws.cell(row=1, column=1, value='Grain Number')
    ws.cell(row=1, column=2, value='Area')
    ws.cell(row=1, column=3, value='Circumference')
    ws.cell(row=1, column=4, value='Length')
    ws.cell(row=1, column=5, value='Width')
    ws.cell(row=1, column=6, value='Diameter')

    # Load the data
    if area:
        for i in range(len(unique)):
            ws.cell(row=i+2, column=1, value=ar[i][0])
            ws.cell(row=i+2, column=2, value=ar[i][1])
    else:
        pass
    if circum:
        for i in range(len(unique)):
            ws.cell(row=i+2, column=3, value=circum[i][1])
    else:
        pass
    if wl:
        for i in range(len(unique)):
            ws.cell(row=i+2, column=4, value=wl[i][1])
            ws.cell(row=i+2, column=5, value=wl[i][2])
            ws.cell(row=i+2, column=6, value=wl[i][3])
    else:
        pass
    
    # Save the data
    wb.save(filename + '.xlsx')
    print('Data saved. Filename: ' + filename + '.xlsx')

