import cv2
import numpy as np
import matplotlib.pyplot as plt
import os
from openpyxl import Workbook

from typing import Dict, List, Optional, Tuple


def surround(array: np.ndarray, coord: Tuple[int, int]):
    '''Takes in an array and a pixel's coordinate. Returns an array that contains the pixel itself and its eight surrounding pixels, in total nine.
    '''
    lis = coord[0]
    ele = coord[1]
    surround = array[lis-1:lis+2, ele-1:ele+2]
    return surround


def area_circumference(image: np.ndarray):
    '''Takes in a labelled marker image. Returns a dictionary that contains the area, circumference and area/circumference ratio of each grain.
    '''
    blank = np.zeros(
        image.shape, np.uint8)  # A blank image of the same dimension as the mock image. Used for labelling
    labels = np.unique(image)[2:]
    # Get a certain label (positive integer that represent one segmented region)
    for label_no in labels:
        result = np.where(image == label_no)  # Exclude other labels
        # Obtain the coordinates of each label
        coordinates = list(zip(result[0], result[1]))
        for coord in coordinates:
            neighbours = surround(image, coord)
            # Test if the surrounding pixels contain 0, the background.
            if -1 in neighbours:
                # If so, put a mark on the blank label image
                blank[coord] = label_no

    region, circumference = np.unique(blank, return_counts=True)
    circumference = circumference[1:]
    region, area = np.unique(image, return_counts=True)
    region = region[2:]
    area = area[2:]
    ratio = area/circumference
    data = list(zip(region, area, circumference, ratio))

    data_l = []
    for i in range(0, len(data)):
        data_l.append(data[i])
    return data_l


def width_length_ellipse(image: np.ndarray, label, visualise = False):
    '''Takes in a labelled marker image and a specific label. Returns the minor- and major-axis length of an ellipse that will fit the grain with that label.
    '''
    blank = np.zeros(image.shape, np.uint8)  # A blank image of the same dimension as the mock image. Used for extraction grain pixels of one label.
    result = np.where(image == label)  # Obtain the coordinates of each label

    coordinates = list(zip(result[0], result[1]))
    for coord in coordinates:
        blank[coord] = 2
    contours, hierarchy = cv2.findContours(blank, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
    cnt = contours[0]
    ellipse = cv2.fitEllipse(cnt)  # Get data about the geometry of the ellipse
    blank = cv2.ellipse(blank, ellipse, 1, 2)  # Visualise
    
    if visualise == False:
        return ellipse[1]
    else:
        return blank


def width_length_rectangle(image: np.ndarray, label, visualise = False):
    blank = np.zeros(image.shape, np.uint8)  # A blank image of the same dimension as the mock image. Used for extraction grain pixels of one label.
    result = np.where(image == label)  # Obtain the coordinates of each label

    coordinates = list(zip(result[0], result[1]))
    for coord in coordinates:
        blank[coord] = 1
    contours, hierarchy = cv2.findContours(blank, 1, 2)
    cnt = contours[0]
    rect = cv2.minAreaRect(cnt)
    box = cv2.boxPoints(rect)
    box = np.int0(box)
    blank = cv2.drawContours(blank, [box], 0, (0, 0, 255), 2)
    
    if visualise == False:
        return rect[1]
    else:
        return blank


def width_length_size(image: np.ndarray):
    unique = np.unique(image)[2:]

    data_l = []

    for i in unique:
        try:
            data = width_length_ellipse(image, i)[0]
            data_l.append((i, data[0], data[1], (data[0]+data[1])/2))
        except:
            pass
    
    return data_l


def data_extraction(image: np.ndarray, filename):
    wb = Workbook()
    ws = wb.active
        
    ac = area_circumference(image)
    wl = width_length_size(image)
    unique = np.unique(image)[2:]

    for i in range(len(unique)):
        ws.cell(row=i+1, column=1, value=ac[i][0])
        ws.cell(row=i+1, column=2, value=ac[i][1])
        ws.cell(row=i+1, column=3, value=ac[i][2])
        ws.cell(row=i+1, column=4, value=ac[i][3])
        ws.cell(row=i+1, column=5, value=wl[i][1])
        ws.cell(row=i+1, column=6, value=wl[i][2])
        ws.cell(row=i+1, column=7, value=wl[i][3])
        
    wb.save(filename)
    return 'Data saved. Filename:' + filename


# Loading actual marker image
image = np.load(
    r'C:\Users\Xin Wenkang\Documents\Scripts\IPHC\Pics\Data extraction\Marker_IHPC_merged.npy')

fig, axe = plt.subplots(1, 1, figsize=(15,15))
axe.imshow(image)
plt.show()

for i in range(max(np.unique(image))):
    try:
        ellipse = width_length_ellipse(image, i, visualise=True)

        fig, axe = plt.subplots(1, 1, figsize=(15,15))
        axe.imshow(ellipse)
        plt.savefig('ellipse (%d).png' % i, dpi=300)
    except:
        pass